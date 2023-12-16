import pandas as pd
from scipy.stats import spearmanr
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

# Load the dataset and prepare to analyze other sheets
file_path = 'data/datasets/microbiota_trustable.xlsx'
sheets = ['Pylum-level microbiota', 'Family-level microbiota', 'Genus-level microbiota']
top_correlations_fertile = {"Pylum": [], "Family": [], "Genus": []}
top_correlations_infertile = {"Pylum": [], "Family": [], "Genus": []}

# Function to calculate correlations
def calculate_correlations(data, top_n=5):
    fertility_status = data['Fertility']
    correlations = {}
    for taxon in data.columns[2:]:  # Start processing data from the third column
        correlation, _ = spearmanr(data[taxon], fertility_status)
        correlations[taxon] = correlation

    # Sort taxa by the absolute value of their correlation coefficient
    sorted_taxa = sorted(correlations.items(), key=lambda x: abs(x[1]), reverse=True)

    # Split into fertile and infertile groups
    fertile_taxa = [taxa for taxa in sorted_taxa if taxa[1] > 0][:top_n]
    infertile_taxa = [taxa for taxa in sorted_taxa if taxa[1] < 0][:top_n]

    return fertile_taxa, infertile_taxa

# Process each sheet
for sheet in sheets:
    # Load each sheet
    data = pd.read_excel(file_path, sheet_name=sheet)

    # Calculate correlations
    top_fertile, top_infertile = calculate_correlations(data)
    level = sheet.split('-')[0]  # Extract level (Pylum, Family, Genus)
    top_correlations_fertile[level].extend(top_fertile)
    top_correlations_infertile[level].extend(top_infertile)

# Function to rearrange data for the Excel output
def rearrange_data_for_excel(correlations):
    max_length = max(len(correlations["Pylum"]), len(correlations["Family"]), len(correlations["Genus"]))
    output_data = []

    for i in range(max_length):
        row = []
        for level in ["Pylum", "Family", "Genus"]:
            if i < len(correlations[level]):
                row.extend([correlations[level][i][0], correlations[level][i][1], ""])
            else:
                row.extend(["", "", ""])
        output_data.append(row[:-1])  # Remove last space column

    return pd.DataFrame(output_data, columns=["Pylum Top", "Correlation Data", "",
                                              "Family Top", "Correlation Data", "",
                                              "Genus Top", "Correlation Data"])

# Function to save the DataFrame in an Excel file with formatting
def save_excel_with_format(df, filename):
    # Create an Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Convert the DataFrame to rows for Excel and add them to the sheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
            if c_idx % 3 == 0:  # Space columns
                ws.column_dimensions[openpyxl.utils.get_column_letter(c_idx)].width = 2

    # Adjust column width based on the content
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[openpyxl.utils.get_column_letter(column_cells[0].column)].width = length

    # Save the workbook
    wb.save(filename)

# Create DataFrames
fertile_df = rearrange_data_for_excel(top_correlations_fertile)
infertile_df = rearrange_data_for_excel(top_correlations_infertile)

# Save the DataFrames to Excel files with formatting
save_excel_with_format(fertile_df, 'data/outputs/fertile_correlations.xlsx')
save_excel_with_format(infertile_df, 'data/outputs/infertile_correlations.xlsx')


